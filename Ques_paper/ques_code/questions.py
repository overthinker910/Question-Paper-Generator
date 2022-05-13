from operator import index
import string, xlrd
import pandas, random
from fpdf import FPDF
import openpyxl

class PDF(FPDF):
    def header(self):
        #logo
        self.image('static/lion.jpg', 10, 8, 25)
        #font
        self.set_font('helvetica', 'B', 20)
        #title
        self.cell(0, 10, 'Question Paper', ln=1, align='C')
        #line-breaks
        self.ln(20)

# pdf = PDF('P', 'mm', 'Letter')



def generate_ques(pathToQues):
    #loading the workbook in wb
    #!!!!add your path for the excel sheet here!!!!
    #SHREYA'S PATH: C:\\Users\\shrey\\Desktop\\question paper\\Question-Paper-Generator-1\\questions\\
    wb = openpyxl.load_workbook(r'C:\\Users\\gaura\\Desktop\\Question-Paper-Generator\\questions\\'+f'{pathToQues}.xlsx')
    #"questions/aoa_excel.xlsx"
    #cheching the sheets present in that workbook
    sheets = wb.sheetnames
    #print(sheets)

    #storing the Sheet1 in sheet1 to print column values
    sh1 = wb['Sheet1']

    #a simple way to print any value inside the cell
    #without any problem
    # data = sh1['A2'].value
    # print(type(data))

    #generating the pdf
    pdf = PDF('P', 'mm', 'A3')
    pdf1 = PDF('P', 'mm', 'A3')
    #adding a page
    pdf.add_page()
    pdf1.add_page()

    #specifying fonts
    pdf.set_font('helvetica', '', 16)
    pdf1.set_font('helvetica', '', 16)
    
    #looping it
    questions_index = []
    no_of_questions=0
    i=1
    j=1
    while no_of_questions<7:
        column_value = 'A'
        answer_column = 'B'
        row_value = random.randint(2, 26)
        # new_column = sh1
        answer_value = column_value + str(row_value)
        answer_value1 = answer_column + str(row_value)
        
        answer_value = str(answer_value)
        answer_value1 = str(answer_value1)

        #appending each question loction in this array for future use

        #printing questions
        data = sh1["{}".format(answer_value)].value
        print(str(i) + '. '+data)
        i=i+1
        no_of_questions=no_of_questions+1

        pdf.cell(40, 30, f'{i-1}'+'. '+data, ln=True)

        #printing answers
        data1 = sh1["{}".format(answer_value1)].value
        print(str(j) + '. '+data1)
        j=j+1
        # no_of_questions=no_of_questions+1

        pdf1.cell(40, 30, f'{j-1}'+'. '+data1, ln=True)
        questions_index.append(row_value)

    pdf.output('pdf_1.pdf')
    pdf1.output('pdf_2.pdf')